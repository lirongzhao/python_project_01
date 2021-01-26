"""
代码做接口自动化测试，基本流程：
1.接口文档--接口测试用例--excel里
2.自动化excel表格读取测试用例数据--得到数据
2.执行测试--requests去执行--执行结果
3.执行结果和预期结果对比 --是否通过 Pass/Failed
4.执行结果写入excel表格中去

封装函数步骤：
1.实现功能的代码写出来
2.参数化 --可变的数据--形参
3.返回值 --函数有没有需要给别人使用

读取数据：第三方库 openpyxl --excel表格数据操作  1）读取  2）写入
1.安装
2.导入 import
excel 表格对象
1.文档---加载进来--工作簿对象--  load_workbook --最好放到同级目录 路径
2.表单对象：
3.单元格对象
注意：写入生效的话-- 一定要保存文件,保存之前一定要关闭excel文件

面试--什么场景用到字典？？？
从excel表中取数据 每一条数据存到一个字典里；最后把每条字典扩充到列表中
"""
import requests
import pprint
import jsonpath
from openpyxl import load_workbook  # 导入库的指定模块，节省资源
# # 加载工作簿
# wb = load_workbook("test_case_api.xlsx")
# # 获取表单
# sheet = wb["register"]
# cell = sheet.cell(row=1,column=1)
# # 获取单元格值--  对它重新赋值--写入
# cell.value='用例编号'
# print(cell.value)
# # 保存/另存为 工作簿
# wb.save("test_case_api.xlsx")

# 读取表格数据函数
def read_data(filename,sheetname):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    # 自动获取最大行号 列号‘  sheet.max_row sheet.max_column
    cases = []  # 存储所有用例
    max_row = sheet.max_row
    for i in range(2,max_row+1):  # range函数取头不取尾
        case = dict(
            case_id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,
            data=sheet.cell(row=i, column=6).value,
            expected_result=sheet.cell(row=i, column=7).value)
        cases.append(case)
    return cases

# 写入数据函数
def write_result(filename,sheetname,final_result,row,column):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value = final_result
    wb.save(filename)

# 发送接口请求的函数
def api_reg_func(url_api,api_data):
    head_reg = {"X-Lemonban-Media-Type": "lemonban.v2"}
    response_reg = requests.post(url=url_api, json=api_data, headers=head_reg)
    return response_reg.json()

#  主函数只在当前文件执行，其他文件导入这个文件也不会运行
if __name__ == '__main__':
    re = read_data('test_case_api.xlsx', 'register')
    print(re)

